#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""期待されるプロセス × ATC学習内容 の対応表を xlsx で出力する。

シート1「対応表」… 職階ごとに、プロセスの各項目とATC学習項目・ねらいを並べる
シート2「ATC学習項目」… 学習項目の一覧と、どの職階で使うか

使い方: python3 build_atc_xlsx.py  → ATC対応表_福田会計.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from atc_map import ATC_TOPICS, LINKS, VERSION
from levels import LEVELS

OUT = "ATC対応表_福田会計.xlsx"

FONT = "Yu Gothic"
thin = Side(style="thin", color="8C8C8C")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

BANDS = {
    "EP": "1F3864", "P": "2E5597", "SD": "3C69B4", "D": "4472C4",
    "SM": "7B96D4", "M": "9DB3E0", "SS": "B4C6E7", "S": "D9E2F3",
    "T": "F2F2F2",
}
WHITE_TEXT = {"EP", "P", "SD", "D"}

# 学習項目の系統ごとの色
GROUPS = [
    ("選択理論の基礎", "A", "E8EEF7"),
    ("目標達成技術", "B", "F3ECDD"),
    ("対人関係・組織", "C", "E9F0E8"),
]


def group_of(code):
    for name, prefix, fill in GROUPS:
        if code.startswith(prefix):
            return name, fill
    return "", "FFFFFF"


def style(cell, size=9, bold=False, color="000000", fill=None,
          h="left", v="top", wrap=True):
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    cell.border = BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def sheet_map(ws):
    headers = [
        ("記号", 7), ("職階", 12), ("テーマ", 15),
        ("期待されるプロセス", 40), ("ATC学習項目", 34), ("ねらい（つながり）", 46),
    ]
    ws.merge_cells("A1:F1")
    style(ws["A1"], 15, True, h="center", v="center")
    ws["A1"] = f"期待されるプロセス × アチーブメントテクノロジー 対応表　税理士法人福田会計（{VERSION}）"
    ws.row_dimensions[1].height = 28

    for i, (label, w) in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=label)
        style(c, 9, True, "FFFFFF", "404040", h="center", v="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 26

    r = 3
    for lv in LEVELS:
        rows = LINKS.get(lv["sym"], [])
        if not rows:
            continue
        first = r
        band = BANDS[lv["sym"]]
        txt = "FFFFFF" if lv["sym"] in WHITE_TEXT else "000000"

        for proc, codes, why in rows:
            topics = "\n".join(
                f"{c}. {ATC_TOPICS[c][0]}\n　　{ATC_TOPICS[c][1]}" for c in codes
            )
            for col, val in enumerate([None, None, None, proc, topics, why], start=1):
                cell = ws.cell(row=r, column=col)
                if val is not None:
                    cell.value = val
                style(cell)
            lines = max(topics.count("\n") + 1, len(why) // 24 + 1, 3)
            ws.row_dimensions[r].height = max(46, lines * 12.5)
            r += 1

        # 記号・職階・テーマは職階ごとに縦結合
        for col, val in ((1, lv["sym"]), (2, lv["title"]), (3, lv["theme"])):
            if r - 1 > first:
                ws.merge_cells(start_row=first, start_column=col,
                               end_row=r - 1, end_column=col)
            c = ws.cell(row=first, column=col, value=val)
            style(c, 10, col <= 2, txt, band, h="center", v="center")

    ws.freeze_panes = "D3"
    ws.print_title_rows = "1:2"
    return r


def sheet_topics(ws):
    used = {}
    for sym, rows in LINKS.items():
        for _, codes, _ in rows:
            for c in codes:
                used.setdefault(c, []).append(sym)

    headers = [("系統", 14), ("記号", 7), ("学習項目", 24), ("内容", 46), ("使う職階", 24)]
    ws.merge_cells("A1:E1")
    style(ws["A1"], 15, True, h="center", v="center")
    ws["A1"] = f"アチーブメントテクノロジー 学習項目一覧（{VERSION}）"
    ws.row_dimensions[1].height = 28

    for i, (label, w) in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=label)
        style(c, 9, True, "FFFFFF", "404040", h="center", v="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 22

    r = 3
    for code, (name, desc) in ATC_TOPICS.items():
        gname, gfill = group_of(code)
        # 上位→下位の並びで職階を出す
        order = [lv["sym"] for lv in LEVELS]
        syms = sorted(set(used.get(code, [])), key=order.index)
        vals = [gname, code, name, desc, " / ".join(syms) or "―"]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=val)
            if col == 1:
                style(c, 9, True, fill=gfill, h="center", v="center")
            elif col == 2:
                style(c, 9, True, fill=gfill, h="center", v="center")
            else:
                style(c, 9, v="center")
        ws.row_dimensions[r].height = 30
        r += 1

    ws.freeze_panes = "A3"
    ws.print_title_rows = "1:2"


NOTES = [
    "※1.ATCは選択理論心理学（W.グラッサー）と成功哲学（ナポレオン・ヒル）を土台に、「実行力強化」「習慣形成」に特化した講座。300の基礎技術で構成される。",
    "※2.学習項目の並びはATCの講座順ではなく、本表のために系統立てたもの。実際の受講内容と突き合わせて確定させること。",
    "※3.「ねらい」は、その職階のプロセスを回すときに、どのATC技術が効くかを書いたもの。評価面談での問いかけに使う。",
    "※4.必要なインプットとしてのATC受講は現在M（部長）に置いているが、本表のとおり技術自体は全職階で使う。受講時期は要検討。",
]


def main() -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "対応表"
    last = sheet_map(ws1)

    for j, n in enumerate(NOTES):
        r = last + 1 + j
        ws1.merge_cells(f"A{r}:F{r}")
        c = ws1.cell(row=r, column=1, value=n)
        c.font = Font(name=FONT, size=8)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws1.row_dimensions[r].height = 14

    ws1.page_setup.orientation = "landscape"
    ws1.page_setup.paperSize = 9  # A4
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 0
    ws1.sheet_properties.pageSetUpPr.fitToPage = True
    ws1.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)

    ws2 = wb.create_sheet("ATC学習項目")
    sheet_topics(ws2)
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.paperSize = 9
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
