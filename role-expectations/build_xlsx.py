#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""期待職能（税理士法人福田会計版）を A3 横1枚に印刷できる xlsx として出力する。

内容は levels.py から読む。元フォーマットは EMPグループ
「16 （A3印刷＆片袖折り）給与テーブル期待職能」。

使い方: python3 build_xlsx.py  → 期待職能_福田会計.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from levels import LEVELS, NOTES, VERSION, flag_of

OUT = "期待職能_福田会計.xlsx"

HEADERS = [
    ("記号", 7),
    ("職階", 13),
    ("グレード", 10),
    ("期待売上高\n（年間）", 13),
    ("テーマ", 16),
    ("昇格の条件", 30),
    ("期待される役割", 38),
    ("測定指標1\n（関与先担当）", 38),
    ("測定指標2\n（総務・経理）", 28),
    ("期待されるプロセス", 28),
    ("必要なインプット", 24),
]

# 職階ごとの帯色（上位ほど濃い）
BANDS = {
    "EP": "1F3864",
    "P": "2E5597",
    "SD": "3C69B4",
    "D": "4472C4",
    "SM": "7B96D4",
    "M": "9DB3E0",
    "SS": "B4C6E7",
    "S": "D9E2F3",
    "T1": "F2F2F2",
    "―": "FFFFFF",
}
WHITE_TEXT = {"EP", "P", "SD", "D"}

FONT = "Yu Gothic"
thin = Side(style="thin", color="8C8C8C")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def numbered(items):
    """(1)…(2)… の形に整形。印つきは末尾に ※印 を添える。"""
    out = []
    for i, item in enumerate(items, start=1):
        text, flag = flag_of(item)
        out.append(f"({i}){text}" + (f"【{flag}】" if flag else ""))
    return "\n".join(out)


def bulleted(items):
    out = []
    for item in items:
        text, flag = flag_of(item)
        out.append(text + (f"【{flag}】" if flag else ""))
    return "\n".join(out)


def promo_cell(lv):
    head = f"＜{lv['promo_to']}への昇格要件＞\n" if lv["promo_to"] else ""
    return head + bulleted(lv["promo"])


def role_cell(lv):
    head = f"＜{lv['role_tag']}＞\n" if lv["role_tag"] else ""
    return head + numbered(lv["role"])


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "期待職能"

    ncols = len(HEADERS)
    last_col = get_column_letter(ncols)

    # タイトル行
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = f"期待職能　税理士法人福田会計（{VERSION}）"
    t.font = Font(name=FONT, size=16, bold=True)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ヘッダー行
    for i, (label, width) in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=i, value=label)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[2].height = 34

    # 本体
    for r, lv in enumerate(LEVELS, start=3):
        row = [
            lv["sym"],
            lv["title"],
            lv["grades"],
            lv["revenue"],
            lv["theme"],
            promo_cell(lv),
            role_cell(lv),
            numbered(lv["kpi1"]),
            numbered(lv["kpi2"]),
            numbered(lv["process"]),
            "\n".join(lv["input"]),
        ]
        band = BANDS[lv["sym"]]
        for i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.border = BORDER
            if i <= 5:  # 記号〜テーマは帯色
                c.fill = PatternFill("solid", fgColor=band)
                color = "FFFFFF" if lv["sym"] in WHITE_TEXT else "000000"
                c.font = Font(name=FONT, size=10, bold=(i <= 2), color=color)
                c.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            else:
                c.font = Font(name=FONT, size=9)
                c.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
        lines = max(str(v).count("\n") + 1 for v in row)
        ws.row_dimensions[r].height = max(52, lines * 13.5)

    # 注記
    note_start = len(LEVELS) + 4
    for j, note in enumerate(NOTES):
        r = note_start + j
        ws.merge_cells(f"A{r}:{last_col}{r}")
        c = ws.cell(row=r, column=1, value=note)
        c.font = Font(name=FONT, size=8)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 14

    # 印刷設定（A3横・1ページ幅に収める）
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2
    )
    ws.print_title_rows = "1:2"
    ws.freeze_panes = "F3"

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
